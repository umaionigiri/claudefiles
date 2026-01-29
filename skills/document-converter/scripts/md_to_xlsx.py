#!/usr/bin/env python3
"""md_to_xlsx.py - Convert Markdown tables to Excel"""
import re
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


def parse_markdown_tables(md_content):
    """Extract tables from Markdown content."""
    tables = []
    lines = md_content.split('\n')
    current_table = []
    in_table = False
    table_title = None

    for i, line in enumerate(lines):
        # Check for heading before table
        if line.startswith('#') and not in_table:
            table_title = line.lstrip('#').strip()

        # Detect table row
        if '|' in line and line.strip().startswith('|'):
            if not in_table:
                in_table = True
            # Skip separator row
            if not re.match(r'^[\|\-:\s]+$', line):
                cells = [cell.strip() for cell in line.strip('|').split('|')]
                current_table.append(cells)
        else:
            if in_table and current_table:
                tables.append({'title': table_title, 'data': current_table})
                current_table = []
                in_table = False
                table_title = None

    if current_table:
        tables.append({'title': table_title, 'data': current_table})

    return tables


def create_xlsx(tables, output_path):
    """Create Excel file from parsed tables."""
    wb = Workbook()

    header_font = Font(name='Meiryo UI', size=10, bold=True)
    cell_font = Font(name='Meiryo UI', size=10)
    header_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for idx, table_info in enumerate(tables):
        if idx == 0:
            ws = wb.active
            ws.title = table_info['title'][:31] if table_info['title'] else 'Sheet1'
        else:
            title = table_info['title'][:31] if table_info['title'] else f'Sheet{idx+1}'
            ws = wb.create_sheet(title=title)

        for row_idx, row_data in enumerate(table_info['data']):
            for col_idx, cell_value in enumerate(row_data):
                cell = ws.cell(row=row_idx+1, column=col_idx+1, value=cell_value)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical='top')

                if row_idx == 0:
                    cell.font = header_font
                    cell.fill = header_fill
                else:
                    cell.font = cell_font

        # Auto-adjust column width
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

    wb.save(output_path)
    print(f"Created: {output_path}")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python3 md_to_xlsx.py <input.md> <output.xlsx>")
        sys.exit(1)

    with open(sys.argv[1], 'r', encoding='utf-8') as f:
        content = f.read()

    tables = parse_markdown_tables(content)
    if tables:
        create_xlsx(tables, sys.argv[2])
    else:
        print("No tables found in Markdown file.")
