#!/usr/bin/env python3
"""json_to_xlsx.py - Convert JSON to Excel"""
import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


def json_to_xlsx(json_path, output_path):
    """Convert JSON array to Excel."""
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    if not isinstance(data, list) or len(data) == 0:
        print("JSON must be an array of objects.")
        return

    wb = Workbook()
    ws = wb.active

    header_font = Font(name='Meiryo UI', size=10, bold=True)
    cell_font = Font(name='Meiryo UI', size=10)
    header_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers from first object keys
    headers = list(data[0].keys())
    for col_idx, header in enumerate(headers):
        cell = ws.cell(row=1, column=col_idx+1, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    # Data rows
    for row_idx, item in enumerate(data):
        for col_idx, key in enumerate(headers):
            value = item.get(key, '')
            if isinstance(value, (dict, list)):
                value = json.dumps(value, ensure_ascii=False)
            cell = ws.cell(row=row_idx+2, column=col_idx+1, value=value)
            cell.font = cell_font
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Auto-adjust column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(output_path)
    print(f"Created: {output_path}")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python3 json_to_xlsx.py <input.json> <output.xlsx>")
        sys.exit(1)
    json_to_xlsx(sys.argv[1], sys.argv[2])
